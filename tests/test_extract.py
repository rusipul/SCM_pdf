import sys
from unittest.mock import MagicMock

from openpyxl import load_workbook
import extract
from extract import parse_grand_total, parse_shipments
from extract import process_pdf


def test_parse_grand_total_extracts_amount():
    text = (
        "Total Number of Shipments총 발송 건수 29\n"
        "Grand Total총액 KRW 7,834,180.00\n"
        "2026년 FedEx 요금 개정 및 기타 주요 변경 사항\n"
    )
    assert parse_grand_total(text) == 7834180.00


def test_parse_grand_total_returns_none_when_absent():
    text = "Summary by Payment Type 결제 유형별 요약\nTotal 합계 2,348,680.00\n"
    assert parse_grand_total(text) is None


SAMPLE_TWO_SHIPMENTS = """\
선적일자 Ship Date 05/27/2026 발송인Sender 수취인Recipient
AWB 번호 Air Waybill Number 872265794268 SHIN JONG SUN ALBERT SHEN
서비스 종류 Service Type 2P OKINS ELECTRONICS ,LTD GREATEK ELECTRONICS INC.
운임Freight Charges 1,539,900.00
차감Deductions 기본 할인 Base Discount (1,148,300.00)
기타비용Other Charges 유류할증 추가요금 Fuel Surcharge 193,840.00
합계Total 585,440.00
49.50%의 유류할증료가 부과되었습니다.
선적일자 Ship Date 05/28/2026 발송인Sender 수취인Recipient
AWB 번호 Air Waybill Number 872319321541 ARIA DONG KAORI ANEZAKI
합계Total 53,880.00
"""


def test_parse_shipments_extracts_all_fields():
    result = parse_shipments(SAMPLE_TWO_SHIPMENTS)
    assert result == [
        {"ship_date": "05/27/2026", "awb_number": "872265794268", "total": 585440.00},
        {"ship_date": "05/28/2026", "awb_number": "872319321541", "total": 53880.00},
    ]


def test_parse_shipments_returns_empty_list_when_no_matches():
    assert parse_shipments("아무 패턴도 없는 텍스트입니다.") == []


def test_parse_shipments_skips_total_without_preceding_awb():
    text = "합계Total 1,000.00\n"
    assert parse_shipments(text) == []


def test_parse_shipments_extracts_total_when_merged_with_dimension_line():
    text = (
        "선적일자 Ship Date 06/15/2026 발송인Sender 수취인Recipient\n"
        "AWB 번호 Air Waybill Number 872999999999 A B\n"
        "규격Dimension 36x34x5cm 합계Total 41,710.00\n"
    )
    result = parse_shipments(text)
    assert result == [
        {"ship_date": "06/15/2026", "awb_number": "872999999999", "total": 41710.00},
    ]


from extract import build_workbook

SHIPMENTS_FIXTURE = [
    {"ship_date": "05/27/2026", "awb_number": "872265794268", "total": 585440.00},
    {"ship_date": "05/28/2026", "awb_number": "872319321541", "total": 53880.00},
]


def test_build_workbook_writes_header_and_rows():
    wb, extracted_sum = build_workbook(SHIPMENTS_FIXTURE, grand_total=639320.00)
    ws = wb.active
    rows = [[cell.value for cell in row] for row in ws.iter_rows()]
    assert rows[0] == ["선적일자", "AWB번호", "Total 금액"]
    assert rows[1] == ["05/27/2026", "872265794268", 585440.00]
    assert rows[2] == ["05/28/2026", "872319321541", 53880.00]
    assert rows[3] == ["합계", "", 639320.00]
    assert extracted_sum == 639320.00


def test_build_workbook_flags_mismatch():
    wb, extracted_sum = build_workbook(SHIPMENTS_FIXTURE, grand_total=999999.00)
    ws = wb.active
    last_row = [cell.value for cell in list(ws.iter_rows())[-1]]
    assert "불일치" in last_row[0]


def test_build_workbook_no_warning_when_grand_total_missing():
    wb, extracted_sum = build_workbook(SHIPMENTS_FIXTURE, grand_total=None)
    ws = wb.active
    rows = [[cell.value for cell in row] for row in ws.iter_rows()]
    assert len(rows) == 4  # header + 2 shipments + sum row, no warning row


def test_process_pdf_writes_xlsx_next_to_pdf(tmp_path, monkeypatch):
    pdf_path = tmp_path / "청구서.pdf"
    pdf_path.write_bytes(b"%PDF-fake")
    monkeypatch.setattr(extract, "extract_pdf_text", lambda path: SAMPLE_TWO_SHIPMENTS)

    result = process_pdf(pdf_path)

    output_path = tmp_path / "청구서.xlsx"
    assert result["output_path"] == output_path
    assert output_path.exists()
    assert result["shipment_count"] == 2
    assert result["extracted_sum"] == 639320.00

    wb = load_workbook(output_path)
    ws = wb.active
    rows = [[cell.value for cell in row] for row in ws.iter_rows()]
    assert rows[0] == ["선적일자", "AWB번호", "Total 금액"]


def test_process_pdf_matches_when_grand_total_agrees(tmp_path, monkeypatch):
    pdf_path = tmp_path / "청구서.pdf"
    pdf_path.write_bytes(b"%PDF-fake")
    text_with_grand_total = SAMPLE_TWO_SHIPMENTS + "\nGrand Total총액 KRW 639,320.00\n"
    monkeypatch.setattr(extract, "extract_pdf_text", lambda path: text_with_grand_total)

    result = process_pdf(pdf_path)

    assert result["matched"] is True
    assert result["grand_total"] == 639320.00


def test_process_pdf_returns_false_matched_when_grand_total_disagrees(tmp_path, monkeypatch):
    pdf_path = tmp_path / "청구서.pdf"
    pdf_path.write_bytes(b"%PDF-fake")
    text_with_wrong_grand_total = SAMPLE_TWO_SHIPMENTS + "\nGrand Total총액 KRW 999,999.00\n"
    monkeypatch.setattr(extract, "extract_pdf_text", lambda path: text_with_wrong_grand_total)

    result = process_pdf(pdf_path)

    assert result["matched"] is False
    assert result["grand_total"] == 999999.00


def test_process_pdf_returns_none_matched_when_grand_total_missing(tmp_path, monkeypatch):
    pdf_path = tmp_path / "청구서.pdf"
    pdf_path.write_bytes(b"%PDF-fake")
    monkeypatch.setattr(extract, "extract_pdf_text", lambda path: SAMPLE_TWO_SHIPMENTS)

    result = process_pdf(pdf_path)

    assert result["grand_total"] is None
    assert result["matched"] is None


def test_process_pdf_skips_file_when_no_shipments_found(tmp_path, monkeypatch):
    pdf_path = tmp_path / "빈청구서.pdf"
    pdf_path.write_bytes(b"%PDF-fake")
    monkeypatch.setattr(extract, "extract_pdf_text", lambda path: "관련 없는 텍스트")

    result = process_pdf(pdf_path)

    assert result["shipment_count"] == 0
    assert result["output_path"] is None
    assert not (tmp_path / "빈청구서.xlsx").exists()


from extract import main


def test_main_skips_non_pdf_files(tmp_path, capsys):
    txt_path = tmp_path / "메모.txt"
    txt_path.write_text("이건 PDF가 아님")

    main([str(txt_path)])

    captured = capsys.readouterr()
    assert "PDF 파일이 아닙니다" in captured.out


def test_main_reports_missing_file(capsys):
    main(["존재하지않는파일.pdf"])

    captured = capsys.readouterr()
    assert "찾을 수 없습니다" in captured.out


def test_main_prints_usage_when_no_args(capsys):
    main([])

    captured = capsys.readouterr()
    assert "사용법" in captured.out


def test_main_processes_multiple_valid_pdfs_independently(tmp_path, monkeypatch, capsys):
    first_pdf = tmp_path / "첫번째.pdf"
    first_pdf.write_bytes(b"%PDF-fake")
    second_pdf = tmp_path / "두번째.pdf"
    second_pdf.write_bytes(b"%PDF-fake")

    monkeypatch.setattr(extract, "extract_pdf_text", lambda path: SAMPLE_TWO_SHIPMENTS)

    main([str(first_pdf), str(second_pdf)])

    assert (tmp_path / "첫번째.xlsx").exists()
    assert (tmp_path / "두번째.xlsx").exists()
    captured = capsys.readouterr()
    assert "첫번째.xlsx" in captured.out
    assert "두번째.xlsx" in captured.out


def test_main_prints_success_summary(tmp_path, monkeypatch, capsys):
    pdf_path = tmp_path / "청구서.pdf"
    pdf_path.write_bytes(b"%PDF-fake")
    monkeypatch.setattr(extract, "extract_pdf_text", lambda path: SAMPLE_TWO_SHIPMENTS)

    main([str(pdf_path)])

    captured = capsys.readouterr()
    assert "청구서.xlsx" in captured.out
    assert "2건" in captured.out


def test_main_warns_when_zero_shipments_found(tmp_path, monkeypatch, capsys):
    pdf_path = tmp_path / "빈청구서.pdf"
    pdf_path.write_bytes(b"%PDF-fake")
    monkeypatch.setattr(extract, "extract_pdf_text", lambda path: "관련 없는 텍스트")

    main([str(pdf_path)])

    captured = capsys.readouterr()
    assert "추출된 건이 없습니다" in captured.out


def test_main_shows_skipped_validation_when_grand_total_missing(tmp_path, monkeypatch, capsys):
    pdf_path = tmp_path / "청구서.pdf"
    pdf_path.write_bytes(b"%PDF-fake")
    monkeypatch.setattr(extract, "extract_pdf_text", lambda path: SAMPLE_TWO_SHIPMENTS)

    main([str(pdf_path)])

    captured = capsys.readouterr()
    assert "검증 생략" in captured.out


def test_main_catches_per_file_exception_and_continues(tmp_path, monkeypatch, capsys):
    pdf_path = tmp_path / "손상된청구서.pdf"
    pdf_path.write_bytes(b"%PDF-fake")

    def raise_error(path):
        raise ValueError("손상된 PDF입니다")

    monkeypatch.setattr(extract, "extract_pdf_text", raise_error)

    main([str(pdf_path)])

    captured = capsys.readouterr()
    assert "처리 중 문제가 발생했습니다" in captured.out
    assert "손상된 PDF입니다" in captured.out


def test_main_continues_processing_remaining_files_after_one_fails(tmp_path, monkeypatch, capsys):
    bad_txt = tmp_path / "메모.txt"
    bad_txt.write_text("이건 PDF가 아님")

    good_pdf = tmp_path / "청구서.pdf"
    good_pdf.write_bytes(b"%PDF-fake")

    monkeypatch.setattr(extract, "extract_pdf_text", lambda path: SAMPLE_TWO_SHIPMENTS)

    main([str(bad_txt), str(good_pdf)])

    captured = capsys.readouterr()
    assert "PDF 파일이 아닙니다" in captured.out
    assert "청구서.xlsx" in captured.out


def test_main_processes_next_file_after_one_raises_exception(tmp_path, monkeypatch, capsys):
    bad_pdf = tmp_path / "손상된청구서.pdf"
    bad_pdf.write_bytes(b"%PDF-fake")

    good_pdf = tmp_path / "청구서.pdf"
    good_pdf.write_bytes(b"%PDF-fake")

    def extract_text_or_raise(path):
        if path == bad_pdf:
            raise ValueError("손상된 PDF입니다")
        return SAMPLE_TWO_SHIPMENTS

    monkeypatch.setattr(extract, "extract_pdf_text", extract_text_or_raise)

    main([str(bad_pdf), str(good_pdf)])

    captured = capsys.readouterr()
    assert "처리 중 문제가 발생했습니다" in captured.out
    assert "청구서.xlsx" in captured.out


def test_main_reconfigures_stdout_when_not_utf8(monkeypatch):
    fake_stdout = MagicMock()
    fake_stdout.encoding = "cp949"
    monkeypatch.setattr(sys, "stdout", fake_stdout)

    main([])

    fake_stdout.reconfigure.assert_called_once_with(encoding="utf-8", errors="replace")


def test_main_reconfigures_stdout_when_encoding_is_none(monkeypatch):
    fake_stdout = MagicMock()
    fake_stdout.encoding = None
    monkeypatch.setattr(sys, "stdout", fake_stdout)

    main([])

    fake_stdout.reconfigure.assert_called_once_with(encoding="utf-8", errors="replace")


def test_main_does_not_reconfigure_stdout_when_already_utf8(monkeypatch):
    fake_stdout = MagicMock()
    fake_stdout.encoding = "utf-8"
    monkeypatch.setattr(sys, "stdout", fake_stdout)

    main([])

    fake_stdout.reconfigure.assert_not_called()


def test_main_prints_mismatch_status_without_crashing(tmp_path, monkeypatch, capsys):
    pdf_path = tmp_path / "청구서.pdf"
    pdf_path.write_bytes(b"%PDF-fake")
    text_with_wrong_grand_total = SAMPLE_TWO_SHIPMENTS + "\nGrand Total총액 KRW 999,999.00\n"
    monkeypatch.setattr(extract, "extract_pdf_text", lambda path: text_with_wrong_grand_total)

    main([str(pdf_path)])

    captured = capsys.readouterr()
    assert "불일치" in captured.out
